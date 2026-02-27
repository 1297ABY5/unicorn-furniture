#!/usr/bin/env python3
"""
UNICORN FURNITURE — ALIEXPRESS PRODUCT IMPORTER
================================================
Connects to AliExpress Affiliate API → Pulls furniture products → 
Populates spreadsheet → Generates site.

SETUP (one-time, 10 minutes):
  1. Go to https://portals.aliexpress.com/ 
  2. Sign up for AliExpress Affiliate Program
  3. Go to https://openservice.aliexpress.com/ 
  4. Apply for "Affiliate API" access (approved in 1-2 days)
  5. Get your APP_KEY, APP_SECRET, and TRACKING_ID
  6. Set them below or as environment variables

USAGE:
  python aliexpress_import.py                    # Interactive mode
  python aliexpress_import.py "luxury bed frame" # Search specific keyword
  python aliexpress_import.py --category beds    # Search by category
  python aliexpress_import.py --bulk             # Bulk import all categories
"""

import os
import sys
import json
import time
import hashlib
import hmac
from datetime import datetime
from pathlib import Path

# ─── TRY OFFICIAL SDK FIRST, FALLBACK TO DIRECT API ─────────────────────
try:
    from aliexpress_api import AliexpressApi, models as ae_models
    HAS_SDK = True
except ImportError:
    HAS_SDK = False

import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════
# CONFIGURATION — Fill these in or set as environment variables
# ═══════════════════════════════════════════════════════════════
APP_KEY = os.environ.get("AE_APP_KEY", "YOUR_APP_KEY_HERE")
APP_SECRET = os.environ.get("AE_APP_SECRET", "YOUR_APP_SECRET_HERE")
TRACKING_ID = os.environ.get("AE_TRACKING_ID", "YOUR_TRACKING_ID_HERE")

# Product markup (your selling price = AliExpress price × MARKUP)
MARKUP_MULTIPLIER = 2.5  # 2.5x = 60% margin
CURRENCY = "USD"  # AliExpress API returns USD, we convert to AED
USD_TO_AED = 3.67

# Output files
SPREADSHEET_OUTPUT = "unicorn-furniture-products.xlsx"
SITE_OUTPUT = "unicorn-furniture-generated.jsx"

# ═══════════════════════════════════════════════════════════════
# FURNITURE CATEGORIES TO SEARCH
# ═══════════════════════════════════════════════════════════════
FURNITURE_SEARCHES = {
    "beds": [
        "luxury modern bed frame velvet",
        "upholstered platform bed king",
        "storage bed hydraulic modern",
        "wall panel headboard bed",
    ],
    "sofas": [
        "luxury sectional sofa L shape",
        "modern bouclé sofa living room",
        "velvet sofa 3 seater modern",
        "modular sofa set living room",
    ],
    "dining": [
        "marble dining table modern luxury",
        "dining table set 6 seater",
        "round dining table modern",
    ],
    "chairs": [
        "accent chair velvet modern",
        "luxury armchair living room",
        "wing chair modern design",
    ],
    "tv": [
        "TV cabinet modern LED light",
        "floating TV stand wall mount",
        "media console modern luxury",
    ],
    "wardrobes": [
        "modern wardrobe sliding door",
        "walk in closet system",
    ],
    "tables": [
        "coffee table modern glass",
        "marble coffee table gold",
        "side table modern luxury",
    ],
    "nightstands": [
        "nightstand modern bedroom",
        "bedside table luxury design",
    ],
}


# ═══════════════════════════════════════════════════════════════
# ALIEXPRESS API CLIENT (Direct API if SDK not available)
# ═══════════════════════════════════════════════════════════════

class AliExpressClient:
    """Direct API client for AliExpress Open Platform."""
    
    BASE_URL = "https://api-sg.aliexpress.com/sync"
    
    def __init__(self, app_key, app_secret, tracking_id):
        self.app_key = app_key
        self.app_secret = app_secret
        self.tracking_id = tracking_id
    
    def _sign(self, params):
        """Generate API signature."""
        sorted_params = sorted(params.items())
        sign_str = self.app_secret
        for k, v in sorted_params:
            sign_str += f"{k}{v}"
        sign_str += self.app_secret
        return hmac.new(
            self.app_secret.encode(), 
            sign_str.encode(), 
            hashlib.sha256
        ).hexdigest().upper()
    
    def _request(self, method, params):
        """Make API request."""
        sys_params = {
            "app_key": self.app_key,
            "method": method,
            "sign_method": "sha256",
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "v": "2.0",
        }
        all_params = {**sys_params, **params}
        all_params["sign"] = self._sign(all_params)
        
        response = requests.get(self.BASE_URL, params=all_params, timeout=30)
        return response.json()
    
    def search_products(self, keywords, category_id=None, page=1, page_size=20, 
                       sort="SALE_PRICE_ASC", min_price=None, max_price=None):
        """Search for affiliate products."""
        params = {
            "keywords": keywords,
            "page_no": str(page),
            "page_size": str(page_size),
            "sort": sort,
            "target_currency": CURRENCY,
            "target_language": "EN",
            "tracking_id": self.tracking_id,
            "ship_to_country": "AE",  # Ship to UAE
        }
        if category_id:
            params["category_ids"] = str(category_id)
        if min_price:
            params["min_sale_price"] = str(min_price)
        if max_price:
            params["max_sale_price"] = str(max_price)
        
        return self._request(
            "aliexpress.affiliate.product.query", 
            params
        )
    
    def get_product_detail(self, product_ids):
        """Get detailed product info."""
        ids = ",".join(str(pid) for pid in product_ids) if isinstance(product_ids, list) else str(product_ids)
        params = {
            "product_ids": ids,
            "target_currency": CURRENCY,
            "target_language": "EN",
            "tracking_id": self.tracking_id,
            "country": "AE",
        }
        return self._request(
            "aliexpress.affiliate.productdetail.get",
            params
        )
    
    def get_hot_products(self, keywords=None, category_id=None, page=1, page_size=20):
        """Get trending/hot products."""
        params = {
            "page_no": str(page),
            "page_size": str(page_size),
            "target_currency": CURRENCY,
            "target_language": "EN",
            "tracking_id": self.tracking_id,
            "ship_to_country": "AE",
        }
        if keywords:
            params["keywords"] = keywords
        if category_id:
            params["category_ids"] = str(category_id)
        
        return self._request(
            "aliexpress.affiliate.hotproduct.query",
            params
        )


def search_with_sdk(keywords, page_size=20):
    """Search using official Python SDK."""
    aliexpress = AliexpressApi(APP_KEY, APP_SECRET, ae_models.Language.EN, ae_models.Currency.USD, TRACKING_ID)
    try:
        products = aliexpress.get_hotproducts(
            keywords=keywords,
            max_sale_price=500,  # USD max price for furniture
            page_size=page_size,
            ship_to_country=ae_models.Country.AE,
            sort=ae_models.SortBy.SALE_PRICE_ASC,
        )
        return products
    except Exception as e:
        print(f"  SDK error: {e}")
        return None


def search_with_direct_api(keywords, page_size=20):
    """Search using direct API client."""
    client = AliExpressClient(APP_KEY, APP_SECRET, TRACKING_ID)
    try:
        result = client.search_products(
            keywords=keywords,
            page_size=page_size,
            sort="SALE_PRICE_ASC",
            min_price=50,   # Min $50 (furniture)
            max_price=2000, # Max $2000
        )
        return result
    except Exception as e:
        print(f"  API error: {e}")
        return None


# ═══════════════════════════════════════════════════════════════
# PRODUCT DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════

def extract_product_data(raw_product, category):
    """Extract and normalize product data from AliExpress API response."""
    
    # Handle both SDK objects and dict responses
    if isinstance(raw_product, dict):
        product_id = raw_product.get("product_id", "")
        title = raw_product.get("product_title", "")
        
        # Price handling
        sale_price = raw_product.get("target_sale_price", "0")
        original_price = raw_product.get("target_original_price", "0")
        
        # Images
        main_image = raw_product.get("product_main_image_url", "")
        small_images = raw_product.get("product_small_image_urls", {})
        if isinstance(small_images, dict):
            image_list = small_images.get("string", [])
        elif isinstance(small_images, list):
            image_list = small_images
        else:
            image_list = []
            
        # Other data
        orders = raw_product.get("lastest_volume", 0)
        rating = raw_product.get("evaluate_rate", "0")
        shop_url = raw_product.get("shop_url", "")
        product_url = raw_product.get("product_detail_url", raw_product.get("promotion_link", ""))
    else:
        # SDK object
        product_id = getattr(raw_product, "product_id", "")
        title = getattr(raw_product, "product_title", "")
        sale_price = str(getattr(raw_product, "target_sale_price", "0"))
        original_price = str(getattr(raw_product, "target_original_price", "0"))
        main_image = getattr(raw_product, "product_main_image_url", "")
        small_images = getattr(raw_product, "product_small_image_urls", [])
        image_list = small_images if isinstance(small_images, list) else []
        orders = getattr(raw_product, "lastest_volume", 0)
        rating = str(getattr(raw_product, "evaluate_rate", "0"))
        product_url = getattr(raw_product, "promotion_link", "")
    
    # Parse prices
    try:
        cost_usd = float(sale_price)
    except (ValueError, TypeError):
        cost_usd = 0
    
    try:
        original_usd = float(original_price)
    except (ValueError, TypeError):
        original_usd = 0
    
    # Calculate selling prices in AED with markup
    cost_aed = round(cost_usd * USD_TO_AED)
    selling_price_aed = round(cost_usd * USD_TO_AED * MARKUP_MULTIPLIER / 10) * 10 - 1  # Round to nice number
    
    # Old price (if there's a discount, show original × markup)
    if original_usd > cost_usd * 1.1:
        old_price_aed = round(original_usd * USD_TO_AED * MARKUP_MULTIPLIER / 10) * 10 - 1
    else:
        old_price_aed = None
    
    # Collect all images (main + additional)
    all_images = [main_image] if main_image else []
    for img in image_list[:3]:  # Max 3 additional
        if img and img not in all_images:
            all_images.append(img)
    
    # Clean up title (remove common AliExpress junk)
    clean_title = title
    for remove in ["Free Shipping", "Hot Sale", "New Arrival", "2024", "2025", "2026"]:
        clean_title = clean_title.replace(remove, "").strip()
    # Capitalize properly
    clean_title = " ".join(w.capitalize() for w in clean_title.split())
    # Truncate if too long
    if len(clean_title) > 60:
        clean_title = clean_title[:57] + "..."
    
    # Determine badge
    badge = ""
    if orders and int(orders) > 100:
        badge = "Best Seller"
    elif old_price_aed:
        badge = "Sale"
    
    return {
        "product_id": f"AE-{product_id}",
        "name": clean_title,
        "category": category,
        "price_aed": selling_price_aed,
        "old_price_aed": old_price_aed,
        "badge": badge,
        "description": "",  # AliExpress titles are descriptive enough
        "colors": "",  # Would need variant API call
        "sizes": "",
        "images": all_images,
        "cost_aed": cost_aed,
        "cost_usd": cost_usd,
        "ae_url": product_url,
        "ae_orders": orders,
        "ae_rating": rating,
        "delivery_days": 21,  # Standard AliExpress to UAE
        "featured": "NO",
        "active": "YES",
    }


# ═══════════════════════════════════════════════════════════════
# SPREADSHEET GENERATION
# ═══════════════════════════════════════════════════════════════

def products_to_spreadsheet(products, output_path):
    """Write products to the Unicorn Furniture spreadsheet format."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    data_font = Font(size=10, name="Arial")
    thin_border = Border(
        left=Side(style='thin', color='E0DCD5'),
        right=Side(style='thin', color='E0DCD5'),
        top=Side(style='thin', color='E0DCD5'),
        bottom=Side(style='thin', color='E0DCD5')
    )
    
    headers = [
        "product_id", "name", "category", "subcategory", "price_aed",
        "old_price_aed", "badge", "description", "colors", "sizes",
        "image_url_1", "image_url_2", "image_url_3", "image_url_4",
        "supplier", "supplier_sku", "cost_aed", "margin_%",
        "stock_qty", "delivery_days", "featured", "active",
        "ae_url", "ae_orders", "ae_rating"
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, p in enumerate(products, 2):
        images = p.get("images", [])
        row_data = [
            p.get("product_id", ""),
            p.get("name", ""),
            p.get("category", ""),
            "",  # subcategory
            p.get("price_aed", 0),
            p.get("old_price_aed", ""),
            p.get("badge", ""),
            p.get("description", ""),
            p.get("colors", ""),
            p.get("sizes", ""),
            images[0] if len(images) > 0 else "",
            images[1] if len(images) > 1 else "",
            images[2] if len(images) > 2 else "",
            images[3] if len(images) > 3 else "",
            "AliExpress",
            p.get("product_id", ""),
            p.get("cost_aed", 0),
            "",  # margin formula
            0,
            p.get("delivery_days", 21),
            p.get("featured", "NO"),
            p.get("active", "YES"),
            p.get("ae_url", ""),
            p.get("ae_orders", 0),
            p.get("ae_rating", ""),
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
            cell.font = data_font
            cell.border = thin_border
        
        # Margin formula
        ws.cell(row=row_idx, column=18, value=f'=ROUND((E{row_idx}-Q{row_idx})/E{row_idx}*100,1)')
    
    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['K'].width = 50
    ws.column_dimensions['W'].width = 60
    
    ws.auto_filter.ref = f"A1:Y{len(products)+1}"
    ws.freeze_panes = "A2"
    
    # Categories sheet
    ws2 = wb.create_sheet("Categories")
    cat_headers = ["category_id", "name", "display_order", "image_url", "description"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    cats = list(dict.fromkeys([p.get("category", "") for p in products]))
    for i, cat in enumerate(cats, 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=cat.replace("-", " ").title())
        ws2.cell(row=i, column=3, value=i-1)
    
    wb.save(output_path)
    print(f"📊 Spreadsheet saved: {output_path} ({len(products)} products)")


# ═══════════════════════════════════════════════════════════════
# MAIN IMPORT PIPELINE
# ═══════════════════════════════════════════════════════════════

def import_category(category, keywords_list, max_per_keyword=10):
    """Import products for one category."""
    products = []
    
    for keywords in keywords_list:
        print(f"  🔍 Searching: '{keywords}'...")
        
        if HAS_SDK:
            raw = search_with_sdk(keywords, page_size=max_per_keyword)
            if raw and hasattr(raw, 'products'):
                raw_products = raw.products
            else:
                raw_products = []
        else:
            result = search_with_direct_api(keywords, page_size=max_per_keyword)
            if result:
                # Navigate nested response
                resp = result.get("aliexpress_affiliate_product_query_response", {})
                resp_result = resp.get("resp_result", {}).get("result", {})
                raw_products = resp_result.get("products", {}).get("product", [])
            else:
                raw_products = []
        
        for raw_p in raw_products:
            try:
                p = extract_product_data(raw_p, category)
                if p["price_aed"] > 100:  # Filter out cheap items
                    products.append(p)
            except Exception as e:
                print(f"    ⚠️ Skipped product: {e}")
        
        print(f"    Found {len(raw_products)} products")
        time.sleep(1)  # Rate limit respect
    
    return products


def bulk_import():
    """Import all furniture categories."""
    all_products = []
    
    print("🚀 UNICORN FURNITURE — AliExpress Bulk Import")
    print("=" * 50)
    
    for category, keywords_list in FURNITURE_SEARCHES.items():
        print(f"\n📦 Category: {category.upper()}")
        products = import_category(category, keywords_list)
        all_products.extend(products)
        print(f"  ✅ Total for {category}: {len(products)} products")
    
    # Remove duplicates by product_id
    seen = set()
    unique = []
    for p in all_products:
        if p["product_id"] not in seen:
            seen.add(p["product_id"])
            unique.append(p)
    
    print(f"\n{'=' * 50}")
    print(f"📊 Total raw products: {len(unique)}")
    
    # ── PREMIUM CURATION ──
    try:
        from premium_curator import PremiumCurator
        print(f"\n🦄 Running Premium Curation Engine...")
        curator = PremiumCurator()
        curated = curator.curate(unique)
        curator.print_report()
        return curated
    except ImportError:
        print("⚠️ premium_curator.py not found — skipping curation")
        # Fallback: mark top sellers as featured
        for p in sorted(unique, key=lambda x: x.get("ae_orders", 0), reverse=True)[:10]:
            p["featured"] = "YES"
        return unique


def search_import(query, category="uncategorized"):
    """Search and import specific products."""
    print(f"🔍 Searching AliExpress for: '{query}'")
    products = import_category(category, [query], max_per_keyword=20)
    print(f"✅ Found {len(products)} products")
    return products


# ═══════════════════════════════════════════════════════════════
# CLI INTERFACE
# ═══════════════════════════════════════════════════════════════

def check_credentials():
    """Check if API credentials are configured."""
    if APP_KEY == "YOUR_APP_KEY_HERE" or not APP_KEY:
        print("=" * 60)
        print("⚠️  ALIEXPRESS API NOT CONFIGURED")
        print("=" * 60)
        print()
        print("To connect to AliExpress, you need API credentials.")
        print()
        print("SETUP STEPS (10 minutes, one-time):")
        print()
        print("1. Go to https://portals.aliexpress.com/")
        print("   Sign up for AliExpress Affiliate Program")
        print()
        print("2. Go to https://openservice.aliexpress.com/")
        print("   Apply for 'Affiliate API' access")
        print("   (Approved in 1-2 business days)")
        print()
        print("3. Once approved, get your credentials from the dashboard:")
        print("   - APP_KEY")
        print("   - APP_SECRET")  
        print("   - TRACKING_ID")
        print()
        print("4. Set them in this script or as environment variables:")
        print("   export AE_APP_KEY='your_key'")
        print("   export AE_APP_SECRET='your_secret'")
        print("   export AE_TRACKING_ID='your_tracking_id'")
        print()
        print("5. Run this script again — it will auto-import products!")
        print()
        print("=" * 60)
        print()
        print("MEANWHILE: You can still use the product spreadsheet")
        print("template (unicorn-furniture-products.xlsx) to manually")
        print("add products with URLs from AliExpress listings.")
        print()
        return False
    return True


def main():
    if not check_credentials():
        sys.exit(0)
    
    if len(sys.argv) > 1:
        if sys.argv[1] == "--bulk":
            products = bulk_import()
        elif sys.argv[1] == "--category":
            cat = sys.argv[2] if len(sys.argv) > 2 else "beds"
            keywords = FURNITURE_SEARCHES.get(cat, [f"modern luxury {cat}"])
            products = import_category(cat, keywords)
        else:
            query = " ".join(sys.argv[1:])
            products = search_import(query)
    else:
        # Interactive mode
        print("🦄 UNICORN FURNITURE — AliExpress Importer")
        print()
        print("Options:")
        print("  1. Bulk import all categories")
        print("  2. Import specific category")
        print("  3. Search by keyword")
        print()
        choice = input("Choose (1/2/3): ").strip()
        
        if choice == "1":
            products = bulk_import()
        elif choice == "2":
            print(f"Categories: {', '.join(FURNITURE_SEARCHES.keys())}")
            cat = input("Category: ").strip()
            keywords = FURNITURE_SEARCHES.get(cat, [f"modern luxury {cat}"])
            products = import_category(cat, keywords)
        else:
            query = input("Search: ").strip() or "luxury modern bed frame"
            products = search_import(query)
    
    if products:
        # Save to spreadsheet
        products_to_spreadsheet(products, SPREADSHEET_OUTPUT)
        
        # Auto-generate site
        print(f"\n🏗️ Generating site...")
        os.system(f"python generate_site.py {SPREADSHEET_OUTPUT} {SITE_OUTPUT}")
        
        print(f"\n🎉 DONE!")
        print(f"  📊 Products: {SPREADSHEET_OUTPUT}")
        print(f"  🌐 Site: {SITE_OUTPUT}")
        print(f"  📦 {len(products)} products imported")
        print(f"\nNext: Push {SITE_OUTPUT} to your Vercel project")
    else:
        print("No products found. Try different search terms.")


if __name__ == "__main__":
    main()
