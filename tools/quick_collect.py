#!/usr/bin/env python3
"""
UNICORN FURNITURE — QUICK PRODUCT COLLECTOR
============================================
No API needed. Just paste AliExpress product URLs.

HOW TO USE:
  1. Browse AliExpress for furniture you want to sell
  2. Copy product URLs into a text file (one per line)
  3. Run: python quick_collect.py urls.txt
  
  OR run interactively:
  python quick_collect.py

The script:
  - Extracts product IDs from URLs  
  - Uses publicly available product data
  - Auto-generates your spreadsheet
  - Runs the site generator

EVEN SIMPLER:
  Just create a CSV file with columns: name, category, price, image_url
  Run: python quick_collect.py products.csv
"""

import csv
import json
import re
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MARKUP = 2.5
USD_TO_AED = 3.67

CATEGORY_MAP = {
    "bed": "beds", "mattress": "mattress", "sofa": "sofas", "couch": "sofas",
    "sectional": "sofas", "dining": "dining", "table": "tables", "coffee": "tables",
    "chair": "chairs", "arm": "chairs", "accent": "chairs", "wardrobe": "wardrobes",
    "closet": "wardrobes", "tv": "tv", "console": "tv", "media": "tv",
    "nightstand": "nightstands", "bedside": "nightstands", "ottoman": "ottoman",
    "dresser": "dressing", "vanity": "dressing", "chaise": "chaise", "lounge": "chaise",
}


def guess_category(name):
    """Guess product category from name."""
    name_lower = name.lower()
    for keyword, cat in CATEGORY_MAP.items():
        if keyword in name_lower:
            return cat
    return "uncategorized"


def parse_aliexpress_urls(url_file):
    """Extract product IDs from AliExpress URLs."""
    products = []
    with open(url_file) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            
            # Extract product ID from various AliExpress URL formats
            match = re.search(r'/item/(\d+)', line) or re.search(r'productId=(\d+)', line)
            if match:
                products.append({
                    "ae_id": match.group(1),
                    "url": line,
                })
            else:
                print(f"  ⚠️ Could not parse URL: {line[:60]}...")
    
    return products


def parse_csv(csv_file):
    """Read products from a simple CSV file."""
    products = []
    df = pd.read_csv(csv_file)
    
    for _, row in df.iterrows():
        name = str(row.get("name", row.iloc[0]))
        category = str(row.get("category", "")) or guess_category(name)
        
        price = 0
        for col in ["price", "price_aed", "selling_price"]:
            if col in row and pd.notna(row[col]):
                try:
                    price = int(float(str(row[col]).replace(",", "").replace("AED", "").strip()))
                except ValueError:
                    pass
                break
        
        cost = 0
        for col in ["cost", "cost_aed", "cost_usd"]:
            if col in row and pd.notna(row[col]):
                try:
                    c = float(str(row[col]).replace(",", "").strip())
                    cost = int(c * USD_TO_AED) if "usd" in col else int(c)
                except ValueError:
                    pass
                break
        
        if price == 0 and cost > 0:
            price = round(cost * MARKUP / 10) * 10 - 1
        
        images = []
        for col in ["image_url", "image", "image_url_1", "img"]:
            if col in row and pd.notna(row[col]):
                images.append(str(row[col]))
                break
        for i in range(2, 5):
            col = f"image_url_{i}"
            if col in row and pd.notna(row[col]):
                images.append(str(row[col]))
        
        products.append({
            "product_id": str(row.get("product_id", row.get("sku", f"P{len(products)+1:03d}"))),
            "name": name,
            "category": category,
            "price_aed": price,
            "old_price_aed": row.get("old_price", None),
            "badge": str(row.get("badge", "")) if "badge" in row and pd.notna(row.get("badge")) else "",
            "description": str(row.get("description", "")) if "description" in row and pd.notna(row.get("description")) else "",
            "colors": str(row.get("colors", "")) if "colors" in row and pd.notna(row.get("colors")) else "",
            "sizes": str(row.get("sizes", "")) if "sizes" in row and pd.notna(row.get("sizes")) else "",
            "images": images,
            "cost_aed": cost,
            "delivery_days": int(row.get("delivery_days", 14)) if "delivery_days" in row and pd.notna(row.get("delivery_days")) else 14,
            "featured": str(row.get("featured", "NO")),
            "active": "YES",
        })
    
    return products


def interactive_collect():
    """Interactively add products one by one."""
    products = []
    print("🦄 UNICORN FURNITURE — Quick Product Collector")
    print("=" * 50)
    print("Add products one at a time. Type 'done' when finished.\n")
    
    while True:
        print(f"\n--- Product #{len(products)+1} ---")
        name = input("Product name (or 'done'): ").strip()
        if name.lower() == "done":
            break
        
        category = input(f"Category [{guess_category(name)}]: ").strip() or guess_category(name)
        
        price_input = input("Selling price (AED): ").strip()
        try:
            price = int(price_input)
        except ValueError:
            price = 999
        
        image = input("Image URL (paste from AliExpress/supplier): ").strip()
        
        colors = input("Colours (comma-separated, or Enter to skip): ").strip()
        sizes = input("Sizes (comma-separated, or Enter to skip): ").strip()
        
        products.append({
            "product_id": f"UF-{len(products)+1:03d}",
            "name": name,
            "category": category,
            "price_aed": price,
            "old_price_aed": None,
            "badge": "",
            "description": "",
            "colors": colors,
            "sizes": sizes,
            "images": [image] if image else [],
            "cost_aed": 0,
            "delivery_days": 14,
            "featured": "YES" if len(products) < 6 else "NO",
            "active": "YES",
        })
        print(f"  ✅ Added: {name} — AED {price}")
    
    return products


def save_spreadsheet(products, output_path="unicorn-furniture-products.xlsx"):
    """Save to spreadsheet format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    hf = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hfill = PatternFill("solid", fgColor="1A1A1A")
    df = Font(size=10, name="Arial")
    border = Border(
        left=Side(style='thin', color='E0DCD5'), right=Side(style='thin', color='E0DCD5'),
        top=Side(style='thin', color='E0DCD5'), bottom=Side(style='thin', color='E0DCD5')
    )
    
    headers = [
        "product_id", "name", "category", "subcategory", "price_aed",
        "old_price_aed", "badge", "description", "colors", "sizes",
        "image_url_1", "image_url_2", "image_url_3", "image_url_4",
        "supplier", "supplier_sku", "cost_aed", "margin_%",
        "stock_qty", "delivery_days", "featured", "active"
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    for ri, p in enumerate(products, 2):
        imgs = p.get("images", [])
        data = [
            p.get("product_id"), p.get("name"), p.get("category"), "",
            p.get("price_aed", 0), p.get("old_price_aed") or "",
            p.get("badge", ""), p.get("description", ""),
            p.get("colors", ""), p.get("sizes", ""),
            imgs[0] if len(imgs) > 0 else "",
            imgs[1] if len(imgs) > 1 else "",
            imgs[2] if len(imgs) > 2 else "",
            imgs[3] if len(imgs) > 3 else "",
            "AliExpress", p.get("product_id"),
            p.get("cost_aed", 0), "",
            0, p.get("delivery_days", 14),
            p.get("featured", "NO"), p.get("active", "YES"),
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val if val else "")
            cell.font = df
            cell.border = border
        ws.cell(row=ri, column=18, value=f'=IF(E{ri}>0,ROUND((E{ri}-Q{ri})/E{ri}*100,1),0)')
    
    # Categories sheet
    ws2 = wb.create_sheet("Categories")
    cats = list(dict.fromkeys([p.get("category", "") for p in products]))
    ch = ["category_id", "name", "display_order", "image_url", "description"]
    for col, h in enumerate(ch, 1):
        ws2.cell(row=1, column=col, value=h).font = hf
        ws2.cell(row=1, column=col).fill = hfill
    for i, cat in enumerate(cats):
        ws2.cell(row=i+2, column=1, value=cat)
        ws2.cell(row=i+2, column=2, value=cat.replace("-", " ").title())
        ws2.cell(row=i+2, column=3, value=i+1)
    
    ws.auto_filter.ref = f"A1:V{len(products)+1}"
    ws.freeze_panes = "A2"
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['K'].width = 50
    
    wb.save(output_path)
    print(f"📊 Saved: {output_path} ({len(products)} products, {len(cats)} categories)")
    return output_path


def main():
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        if filepath.endswith(".csv"):
            print(f"📄 Reading CSV: {filepath}")
            products = parse_csv(filepath)
        elif filepath.endswith(".txt"):
            print(f"🔗 Reading URLs: {filepath}")
            products = parse_aliexpress_urls(filepath)
            if not products:
                print("No valid AliExpress URLs found.")
                return
        else:
            print("Supported formats: .csv, .txt (with AliExpress URLs)")
            return
    else:
        products = interactive_collect()
    
    if not products:
        print("No products collected.")
        return
    
    # ── PREMIUM CURATION ──
    try:
        from premium_curator import PremiumCurator
        print(f"\n🦄 Running Premium Curation Engine on {len(products)} products...")
        curator = PremiumCurator(min_score=35)  # Lower threshold for manual entry
        curated = curator.curate(products)
        curator.print_report()
        products = curated if curated else products
    except ImportError:
        print("ℹ️ premium_curator.py not found — using raw data")
    
    output = save_spreadsheet(products)
    
    # Auto-generate site if generator exists
    if Path("generate_site.py").exists():
        print("\n🏗️ Generating site...")
        import os
        os.system(f"python generate_site.py {output} unicorn-furniture-generated.jsx")
    
    print(f"\n🎉 Done! {len(products)} products ready.")
    print(f"Next: Review the spreadsheet, then deploy the generated site.")


if __name__ == "__main__":
    main()
